"""命令行入口:python -m fulltext_fetcher [输入...] [选项]

输入可以是 DOI、标题、arXiv id(位置参数,多个),或用 --input-file 从文件逐行读取。
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import List

from .config import DEFAULT_SOURCE_ORDER, Config
from .pipeline import Pipeline


def _read_input_file(path: str) -> List[str]:
    """从文件读取输入。按扩展名分发:
      - .csv          → 识别 doi/title 列(无表头则取每行首个非空单元格)
      - .xlsx/.xlsm   → 同 csv 逻辑(需可选依赖 openpyxl)
      - 其它(.txt)   → 逐行,# 开头为注释
    """
    low = path.lower()
    if low.endswith(".csv"):
        return _read_csv(path)
    if low.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(path)
    return _read_text_lines(path)


def _read_text_lines(path: str) -> List[str]:
    out: List[str] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                out.append(line)
    return out


def _extract_from_rows(rows: List[List[str]]) -> List[str]:
    """从二维表格提取输入:有 doi/title 表头则优先 doi、回退 title;
    否则当作无表头,取每行首个非空(非 # 注释)单元格。"""
    out: List[str] = []
    if not rows:
        return out
    header = [c.strip().lower() for c in rows[0]]
    doi_i = header.index("doi") if "doi" in header else -1
    title_i = header.index("title") if "title" in header else -1
    if doi_i >= 0 or title_i >= 0:
        for r in rows[1:]:
            val = ""
            if 0 <= doi_i < len(r) and r[doi_i].strip():
                val = r[doi_i].strip()
            elif 0 <= title_i < len(r) and r[title_i].strip():
                val = r[title_i].strip()
            if val:
                out.append(val)
    else:
        for r in rows:
            for c in r:
                cell = c.strip()
                if cell and not cell.startswith("#"):
                    out.append(cell)
                    break
    return out


def _read_csv(path: str) -> List[str]:
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = [[(c if c is not None else "") for c in row] for row in csv.reader(f)]
    return _extract_from_rows(rows)


def _read_xlsx(path: str) -> List[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "读取 .xlsx 需要可选依赖 openpyxl:pip install openpyxl;"
            "或把表格另存为 .csv 再用 -f data.csv(csv 走标准库、零额外依赖)。"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = [
            [("" if c.value is None else str(c.value)).strip() for c in row]
            for row in ws.iter_rows()
        ]
    finally:
        wb.close()
    return _extract_from_rows(rows)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="fulltext_fetcher",
        description="输入 DOI/标题/arXiv,全自动榨干全网可及免费全文并下载入库。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "子命令:\n"
            "  scholar    谷歌学术爬虫(避免/自动过人机验证,抓元数据 + 下 PDF)。\n"
            "             用法: python -m fulltext_fetcher scholar \"标题或DOI\" [选项]\n"
            "             详见: python -m fulltext_fetcher.scholar --help\n"
            "\n"
            "不带子命令时即为上面的开放获取(OA)全文抓取主流程。"
        ),
    )
    p.add_argument("inputs", nargs="*", help="DOI / 标题 / arXiv id(可多个)")
    p.add_argument("-f", "--input-file",
                   help="从文件读取输入:.txt 逐行(# 注释)/ .csv / .xlsx 自动识别 doi、title 列")
    p.add_argument("--email", default=os.environ.get("FULLTEXT_EMAIL", ""),
                   help="联系邮箱(Unpaywall 必需真实邮箱;OpenAlex/Crossref 礼貌池)")
    p.add_argument("-o", "--out", default="out", help="输出目录(默认 out)")
    p.add_argument("-c", "--concurrency", type=int, default=4, help="并发输入数(默认 4)")
    p.add_argument("--timeout", type=float, default=30.0, help="单请求超时秒(默认 30)")
    p.add_argument("--max-retries", type=int, default=3, help="429/5xx 重试次数(默认 3)")
    p.add_argument("--per-host-interval", type=float, default=0.34,
                   help="每域最小请求间隔秒(默认 0.34;批量验证可调小到 0.15~0.2)")
    p.add_argument("--oa-only", action="store_true", help="仅尝试直链候选,跳过落地页兜底")
    p.add_argument("--no-download", action="store_true", help="只定位不下载(快速验证源命中)")
    p.add_argument("--no-resume", action="store_true", help="不跳过已成功输入(关闭断点续跑)")
    p.add_argument("--retry-failed", action="store_true",
                   help="续跑时也重试上次的永久失败项(如 403/404;默认只重试超时/5xx 等临时失败)")
    p.add_argument("--print-json", action="store_true",
                   help="stdout 输出 JSON({summary,results}),供父程序/子进程接入")
    p.add_argument("--sources", help=f"逗号分隔的源及顺序(默认全部:{','.join(DEFAULT_SOURCE_ORDER)})")
    p.add_argument("--openalex-key", default=os.environ.get("OPENALEX_KEY"))
    p.add_argument("--s2-key", default=os.environ.get("S2_KEY"))
    p.add_argument("--core-key", default=os.environ.get("CORE_KEY"))
    p.add_argument("--zotero-key", default=os.environ.get("ZOTERO_KEY"),
                   help="Zotero API Key(自备);连同 --zotero-library-id 提供后,"
                        "下载成功的文献自动写入你的 Zotero 库(默认取环境变量 ZOTERO_KEY)")
    p.add_argument("--zotero-library-id", default=os.environ.get("ZOTERO_LIBRARY_ID"),
                   help="Zotero 库 ID:个人库用 userID、群组库用 groupID(默认取环境变量 ZOTERO_LIBRARY_ID)")
    p.add_argument("--zotero-library-type",
                   default=os.environ.get("ZOTERO_LIBRARY_TYPE", "user"),
                   choices=["user", "group"],
                   help="Zotero 库类型:user(个人,默认)或 group(群组)")
    p.add_argument("--snapshot-db", default=os.environ.get("FULLTEXT_SNAPSHOT_DB"),
                   help="本地快照 SQLite 路径(由 python -m fulltext_fetcher.ingest 生成);"
                        "配置后 DOI 先走本地、零额度零限速")
    p.add_argument("--institutional", action="store_true",
                   help="启用机构订阅直链源 publisher_direct(对订阅/混合出版商也构造 PDF 直链)。"
                        "仅供拥有合法机构订阅、对内容有访问权者使用;无订阅时直链会 401/403 被 %PDF 校验过滤")
    # 机构订阅三件套(默认全空 → 行为与未启用逐字节一致;详见 机构订阅集成设计.md)。
    # Cookie/前缀建议走环境变量而非命令行明文,避免留在 shell 历史里。
    p.add_argument("--ezproxy-prefix", default=os.environ.get("EZPROXY_PREFIX"),
                   help="EZproxy 接入点:前缀式如 \"https://login.ezproxy.uni.edu/login?url=\";"
                        "或主机名改写式代理域(裸域名)如 \"ezproxy.uni.edu\"(自动识别两种形式;"
                        "默认取环境变量 EZPROXY_PREFIX)")
    p.add_argument("--institution-cookie", default=os.environ.get("INSTITUTION_COOKIE"),
                   help="机构 SSO/EZproxy 登录后的会话 Cookie 串(\"k1=v1; k2=v2\");"
                        "强烈建议用环境变量 INSTITUTION_COOKIE 传入,不留 shell 历史;绝不入日志/产物")
    p.add_argument("--institution-domain", action="append", default=None,
                   metavar="DOMAIN",
                   help="仅对这些出版商域名启用机构通道(可重复给出,或单值内用逗号分隔,"
                        "如 sciencedirect.com,onlinelibrary.wiley.com);不给=不改写任何域名")
    p.add_argument("--enable-scihub", action="store_true",
                   help="启用 Sci-Hub 兜底(注意:合规风险,默认关闭)")
    p.add_argument("--scihub-mirror", default="https://sci-hub.se")
    p.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    return p


def main(argv: List[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    inputs: List[str] = list(args.inputs)
    if args.input_file:
        inputs.extend(_read_input_file(args.input_file))
    if not inputs:
        print("错误:未提供任何输入。示例:\n"
              '  python -m fulltext_fetcher "10.1038/nature12373" --email you@uni.edu\n'
              '  python -m fulltext_fetcher -f dois.txt --email you@uni.edu', file=sys.stderr)
        return 2

    cfg = Config(
        email=args.email or "anonymous@example.com",
        openalex_key=args.openalex_key,
        s2_key=args.s2_key,
        core_key=args.core_key,
        snapshot_db=args.snapshot_db,
        out_dir=args.out,
        concurrency=args.concurrency,
        timeout=args.timeout,
        max_retries=args.max_retries,
        per_host_interval=args.per_host_interval,
        oa_only=args.oa_only,
        no_download=args.no_download,
        resume=not args.no_resume,
        retry_failed=args.retry_failed,
        institutional=args.institutional,
        ezproxy_prefix=(args.ezproxy_prefix or None),
        institution_cookie=(args.institution_cookie or None),
        institution_domains=[d.strip() for raw in (args.institution_domain or [])
                             for d in raw.split(",") if d.strip()],
        enable_scihub=args.enable_scihub,
        scihub_mirror=args.scihub_mirror,
        log_level=args.log_level,
    )
    if args.sources:
        cfg.sources = [s.strip() for s in args.sources.split(",") if s.strip()]
    # 机构模式:把 publisher_direct 接入源顺序(置于免费 OA 源之后、兜底 websearch 之前);
    # 非机构模式绝不注入,避免每条 DOI 多一条 0 候选的空尝试。该源自身也按 cfg.institutional 二次把关。
    if cfg.institutional and "publisher_direct" not in cfg.sources:
        if "websearch" in cfg.sources:
            cfg.sources.insert(cfg.sources.index("websearch"), "publisher_direct")
        else:
            cfg.sources.append("publisher_direct")

    pipe = Pipeline(cfg)
    if cfg.email_is_placeholder():
        pipe.log.warning("未提供真实邮箱:Unpaywall 可能返回 422,建议加 --email you@uni.edu")
    if cfg.institutional:
        pipe.log.warning("已启用机构订阅直链源 publisher_direct:仅供拥有合法机构订阅、"
                         "对内容有访问权者使用;无订阅的直链会 401/403 被 %PDF 校验过滤。")
    if cfg.enable_scihub:
        pipe.log.warning("注意:已启用 Sci-Hub 兜底,存在合规/法律风险,使用者自负。")

    summary = pipe.run(inputs)

    # 可选:把下载成功的文献直写用户自备的 Zotero 库(仅在提供了 key + library-id 时触发;
    # 全程容错,任何网络错误只记日志,不影响已完成的抓取结果与后续汇总/JSON 输出)。
    if args.zotero_key and args.zotero_library_id:
        from . import zotero
        ok_items = [r for r in pipe.results if getattr(r, "success", False)]
        n_zot = zotero.push_items(args.zotero_key, args.zotero_library_id,
                                  args.zotero_library_type, ok_items, pipe.log)
        summary["zotero_uploaded"] = n_zot

    if args.print_json:
        import json
        payload = {"summary": summary, "results": [r.to_dict() for r in pipe.results]}
        # 强制 UTF-8 字节到 stdout(跨平台/跨进程稳定,父程序按 utf-8 解析即可)
        data = (json.dumps(payload, ensure_ascii=False) + "\n").encode("utf-8")
        try:
            sys.stdout.buffer.write(data)
            sys.stdout.buffer.flush()
        except AttributeError:
            sys.stdout.write(data.decode("utf-8"))
    else:
        print_summary(summary, cfg.out_dir)
    return 0 if summary["success"] > 0 or summary["processed"] == 0 else 1


def print_summary(summary: dict, out_dir: str) -> None:
    print("\n===== fulltext_fetcher 运行汇总 =====")
    print(f"处理 {summary['processed']} 条,成功 {summary['success']},失败 {summary['miss']},"
          f"成功率 {summary['success_rate']*100:.0f}%,用时 {summary['elapsed_sec']}s")
    if summary["by_source"]:
        print("命中来源:", ", ".join(f"{k}={v}" for k, v in summary["by_source"].items()))
    print(f"产物目录:{out_dir}/  (pdfs/ + metadata.jsonl + attempts.jsonl + "
          f"summary.json + results.csv + report.html + run.log)")


def _selftest() -> int:
    """不联网 selftest:验证 CSV / 无表头 / .txt 输入解析。
    触发:python -m fulltext_fetcher.cli --selftest
    """
    import csv as _csv
    import os
    import tempfile

    with tempfile.TemporaryDirectory() as d:
        # ① 有 doi/title 表头:每行优先取 doi,doi 空时回退 title
        p1 = os.path.join(d, "headered.csv")
        with open(p1, "w", newline="", encoding="utf-8-sig") as f:
            w = _csv.writer(f)
            w.writerow(["title", "doi", "year"])
            w.writerow(["Paper One", "10.1000/x", "2020"])
            w.writerow(["Paper Two", "", "2021"])        # doi 空 → 回退 title
            w.writerow(["", "10.2000/z", "2022"])
        got = _read_input_file(p1)
        assert got == ["10.1000/x", "Paper Two", "10.2000/z"], got

        # ② 无表头:取每行首个非空、非 # 注释单元格
        p2 = os.path.join(d, "headerless.csv")
        with open(p2, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(["10.5000/aaa", "ignored"])
            w.writerow(["", "10.6000/bbb"])              # 首列空 → 取第二列
            w.writerow(["# comment", "10.7000/ccc"])     # 首列注释 → 取第二列
        got2 = _read_input_file(p2)
        assert got2 == ["10.5000/aaa", "10.6000/bbb", "10.7000/ccc"], got2

        # ③ .txt 逐行:空行 / # 注释跳过,两端空白裁剪
        p3 = os.path.join(d, "list.txt")
        with open(p3, "w", encoding="utf-8") as f:
            f.write("10.8000/d\n# comment line\n\n  10.9000/e  \n")
        got3 = _read_input_file(p3)
        assert got3 == ["10.8000/d", "10.9000/e"], got3

        # ④ _extract_from_rows 直测:表头大小写 / 空格不敏感,doi 优先、回退 title
        rows = [[" DOI ", "Title"], ["10.3000/p", "T"], ["", "Only Title"]]
        assert _extract_from_rows(rows) == ["10.3000/p", "Only Title"], _extract_from_rows(rows)

    # ⑤ 机构订阅参数解析(离线,不跑主流程):--institution-domain 可重复 + 逗号分隔混用;
    #    默认(不给)→ None;Cookie/前缀默认取环境变量。此处只验证 argparse 层与拆分逻辑。
    parser = build_parser()
    a = parser.parse_args(["10.1/x", "--institutional",
                           "--ezproxy-prefix", "https://login.ezproxy.uni.edu/login?url=",
                           "--institution-cookie", "k1=v1; k2=v2",
                           "--institution-domain", "sciencedirect.com,pubs.acs.org",
                           "--institution-domain", "onlinelibrary.wiley.com"])
    assert a.institutional and a.ezproxy_prefix.endswith("login?url="), a.ezproxy_prefix
    assert a.institution_cookie == "k1=v1; k2=v2"
    domains = [d.strip() for raw in (a.institution_domain or [])
               for d in raw.split(",") if d.strip()]
    assert domains == ["sciencedirect.com", "pubs.acs.org", "onlinelibrary.wiley.com"], domains
    a2 = parser.parse_args(["10.1/x"])
    assert a2.institution_domain is None      # 默认不给 → None → Config 得到空列表
    assert not a2.institutional

    print("CLI_OK")
    return 0


if __name__ == "__main__":
    import sys
    if "--selftest" in sys.argv[1:]:
        raise SystemExit(_selftest())
    raise SystemExit(main())
